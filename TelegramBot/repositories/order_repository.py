from openpyxl import load_workbook, Workbook
from typing import Dict, Any
from models.cart import CartItem


class OrderExcelRepository:
    """Репозиторий для сохранения заказов в Excel."""

    def __init__(self, file_path: str) -> None:
        """Инициализация репозитория с указанием пути к файлу.

        Args:
            file_path: Путь к файлу Excel для хранения заказов.
        """
        self._file_path = file_path
        self._header = [
            "Order ID", "User ID", "Total", "Name", "Address",
            "Phone", "Status", "Items"
        ]

    def save_order(self, order: Dict[str, Any]) -> None:
        """Сохранение заказа в Excel-файл.

        Args:
            order: Данные заказа для сохранения.
        """
        try:
            wb = load_workbook(self._file_path)
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(self._header)

        ws = wb.active
        items_str = "; ".join(
            [f"{item.product.name} x{item.quantity}" for item in order["items"]]
        )
        ws.append([
            order["id"],
            order["user_id"],
            order["total"],
            order["name"],
            order["address"],
            order["phone"],
            order["status"],
            items_str,
        ])
        wb.save(self._file_path)