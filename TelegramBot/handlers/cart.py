# handlers/cart.py
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import Command
import logging
import json
import uuid
import requests
from openpyxl import load_workbook, Workbook
from services.api_provider import api_provider as api
from models.catalog import Product
from models.cart import Cart, CartItem
from redis.asyncio import Redis
from config import REDIS_DSN

router = Router()
logger = logging.getLogger(__name__)

# Инициализация Redis
redis_client = Redis.from_url(REDIS_DSN, decode_responses=True)

class CartStates(StatesGroup):
    viewing_cart = State()
    waiting_for_name = State()
    waiting_for_address = State()
    waiting_for_phone = State()

CART_PREFIX = "cart"
REMOVE_FROM_CART = "remove_from_cart_"
CHECKOUT = "checkout"

async def get_cart(user_id: int) -> Cart:
    """Получение корзины из Redis"""
    cart_data = await redis_client.get(f"cart:{user_id}")
    if cart_data:
        return Cart.from_dict(json.loads(cart_data))
    return Cart(items=[])

async def save_cart(user_id: int, cart: Cart):
    """Сохранение корзины в Redis"""
    await redis_client.set(f"cart:{user_id}", json.dumps(cart.to_dict()))

async def clear_cart(user_id: int):
    """Очистка корзины в Redis"""
    await redis_client.delete(f"cart:{user_id}")

@router.message(Command("cart"))
async def cart_command(message: Message, state: FSMContext):
    await show_cart(message, state, message.from_user.id)

@router.callback_query(F.data == CART_PREFIX)
async def show_cart_callback(callback: CallbackQuery, state: FSMContext):
    await show_cart(callback.message, state, callback.from_user.id)
    await callback.answer()

async def show_cart(message: Message, state: FSMContext, user_id: int):
    cart = await get_cart(user_id)
    if cart.is_empty():
        kb = InlineKeyboardBuilder()
        kb.button(text="📋 Каталог", callback_data="catalog")
        kb.adjust(1)
        await message.answer("🛒 Ваша корзина пуста", reply_markup=kb.as_markup())
        return
    response = "🛒 Ваша корзина:\n"
    kb = InlineKeyboardBuilder()
    for item in cart.items:
        response += f"{item.product.name} x{item.quantity} - {item.product.price * item.quantity} ₽\n"
        kb.button(text=f"Удалить {item.product.name}", callback_data=f"{REMOVE_FROM_CART}{item.product.id}")
    response += f"Итого: {cart.get_total()} ₽"
    kb.button(text="Оформить заказ", callback_data=CHECKOUT)
    kb.button(text="◀️ Назад к каталогу", callback_data="catalog")
    kb.adjust(1)
    await state.set_state(CartStates.viewing_cart)
    await message.answer(response, reply_markup=kb.as_markup())

@router.callback_query(F.data.startswith(REMOVE_FROM_CART))
async def remove_from_cart(callback: CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[1])
    user_id = callback.from_user.id
    cart = await get_cart(user_id)
    cart.remove_item(product_id)
    await save_cart(user_id, cart)
    await callback.answer("Товар удалён из корзины")
    await show_cart(callback.message, state, user_id)

@router.callback_query(F.data == CHECKOUT)
async def start_checkout(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    cart = await get_cart(user_id)
    if cart.is_empty():
        await callback.message.edit_text("🛒 Ваша корзина пуста", reply_markup=back_to_catalog_kb())
        return
    await state.set_state(CartStates.waiting_for_name)
    await callback.message.edit_text("Введите ваше имя:")

@router.message(F.text)  # Убрали state из декоратора
async def process_name(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state != CartStates.waiting_for_name.state:
        return
    await state.update_data(name=message.text)
    await CartStates.waiting_for_address.set()
    await message.answer("Введите ваш адрес:")

@router.message(F.text)  # Убрали state из декоратора
async def process_address(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state != CartStates.waiting_for_address.state:
        return
    await state.update_data(address=message.text)
    await CartStates.waiting_for_phone.set()
    await message.answer("Введите ваш телефон:")

@router.message(F.text)  # Убрали state из декоратора
async def process_phone(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state != CartStates.waiting_for_phone.state:
        return
    user_id = message.from_user.id
    data = await state.get_data()
    cart = await get_cart(user_id)
    total = cart.get_total()
    order = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "items": cart.items[:],  # Копируем список
        "total": total,
        "name": data["name"],
        "address": data["address"],
        "phone": message.text,
        "status": "pending"
    }
    payment = create_payment(total, order["id"])
    save_order_to_excel(order)
    kb = InlineKeyboardBuilder()
    kb.button(text="Оплатить", url=payment["confirmation"]["confirmation_url"])
    kb.button(text="◀️ Назад к каталогу", callback_data="catalog")
    await message.answer(
        f"Заказ сформирован! Сумма: {total} ₽\nДля оплаты перейдите по ссылке:",
        reply_markup=kb.as_markup()
    )
    await clear_cart(user_id)  # Очищаем корзину после оформления
    await state.finish()

def create_payment(amount: float, order_id: str) -> dict:
    url = "https://api.yookassa.ru/v3/payments"
    auth = ("YOUR_SHOP_ID", "YOUR_SECRET_KEY")  # Замените на свои данные YooKassa
    data = {
        "amount": {"value": f"{amount:.2f}", "currency": "RUB"},
        "confirmation": {"type": "redirect", "return_url": "https://t.me/your_bot"},
        "description": f"Оплата заказа {order_id}"
    }
    response = requests.post(url, json=data, auth=auth)
    response.raise_for_status()
    return response.json()

def save_order_to_excel(order: dict):
    file_path = "orders.xlsx"
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Order ID", "User ID", "Total", "Name", "Address", "Phone", "Status", "Items"])
    ws = wb.active
    items_str = "; ".join([f"{item.product.name} x{item.quantity}" for item in order["items"]])
    ws.append([order["id"], order["user_id"], order["total"], order["name"], order["address"], order["phone"], order["status"], items_str])
    wb.save(file_path)

def back_to_catalog_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="📋 Каталог", callback_data="catalog")
    return kb.as_markup()

def register_handlers(dp):
    dp.include_router(router)